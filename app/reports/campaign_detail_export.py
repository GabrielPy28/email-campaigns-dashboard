"""Exportación de resultados detallados de una campaña (Excel / CSV por sección o completo)."""

from __future__ import annotations

import csv
import io
import os
import re
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Any, Literal

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import desc, func
from sqlalchemy.orm import Session, joinedload

from app.db import models
from app.tracking.device_category import classify_device_from_user_agent

Section = Literal[
    "full",
    "resumen",
    "actividad_destinatarios",
    "dispositivos",
    "clics_boton",
    "brevo_interno",
    "remitente",
]

BREVO_EXPORT_NOTE = (
    "Sin datos: clave API Brevo no configurada, error de red o respuesta HTTP de error."
)


def _sanitize_filename(name: str, max_len: int = 80) -> str:
    s = re.sub(r"[^\w\s\-]", "", name or "campana", flags=re.UNICODE)
    s = re.sub(r"\s+", "_", s.strip()) or "campana"
    return s[:max_len]


def _device_for_report_row(
    stored_category: str | None, user_agent: str | None
) -> str:
    if stored_category in ("desktop", "mobile", "tablet", "other"):
        return stored_category
    return classify_device_from_user_agent(user_agent)


def _fetch_campaign(db: Session, cid: uuid.UUID) -> models.Campaign | None:
    return (
        db.query(models.Campaign)
        .options(joinedload(models.Campaign.sender))
        .filter(models.Campaign.id == cid)
        .first()
    )


def _opens_recipients(
    db: Session, cid: uuid.UUID
) -> tuple[int, int, int, list[dict[str, Any]]]:
    total_recipients = (
        db.query(func.count(models.CampaignRecipient.id))
        .filter(models.CampaignRecipient.campaign_id == cid)
        .scalar()
        or 0
    )
    rows = (
        db.query(
            models.CampaignRecipient.recipient_id,
            models.CampaignRecipient.email,
            func.count(models.EmailOpen.id).label("open_count"),
            func.max(models.EmailOpen.opened_at).label("last_open_at"),
        )
        .select_from(models.CampaignRecipient)
        .outerjoin(
            models.EmailOpen,
            (models.EmailOpen.campaign_id == models.CampaignRecipient.campaign_id)
            & (models.EmailOpen.recipient_id == models.CampaignRecipient.recipient_id),
        )
        .filter(models.CampaignRecipient.campaign_id == cid)
        .group_by(models.CampaignRecipient.recipient_id, models.CampaignRecipient.email)
        .order_by(models.CampaignRecipient.email.asc())
        .all()
    )
    recipients = [
        {
            "recipient_id": str(r.recipient_id),
            "email": r.email,
            "open_count": int(r.open_count or 0),
            "last_open_at": r.last_open_at.isoformat() if r.last_open_at else None,
        }
        for r in rows
    ]
    unique_open = sum(1 for x in recipients if x["open_count"] > 0)
    total_opens = sum(x["open_count"] for x in recipients)
    return total_recipients, unique_open, total_opens, recipients


def _clicks_by_recipient_map(db: Session, cid: uuid.UUID) -> dict[str, dict[str, Any]]:
    rows = (
        db.query(
            models.CampaignRecipient.recipient_id,
            models.CampaignRecipient.email,
            func.count(models.EmailClick.id).label("click_count"),
            func.max(models.EmailClick.clicked_at).label("last_click_at"),
        )
        .select_from(models.CampaignRecipient)
        .outerjoin(
            models.EmailClick,
            (models.EmailClick.campaign_id == models.CampaignRecipient.campaign_id)
            & (models.EmailClick.recipient_id == models.CampaignRecipient.recipient_id),
        )
        .filter(models.CampaignRecipient.campaign_id == cid)
        .group_by(models.CampaignRecipient.recipient_id, models.CampaignRecipient.email)
        .all()
    )
    return {
        str(r.recipient_id): {
            "email": r.email,
            "click_count": int(r.click_count or 0),
            "last_click_at": r.last_click_at.isoformat() if r.last_click_at else None,
        }
        for r in rows
    }


def _clicks_by_button_rows(db: Session, cid: uuid.UUID) -> list[dict[str, Any]]:
    rows = (
        db.query(
            models.EmailClick.button_id,
            func.count(models.EmailClick.id).label("cnt"),
        )
        .filter(models.EmailClick.campaign_id == cid)
        .group_by(models.EmailClick.button_id)
        .order_by(func.count(models.EmailClick.id).desc())
        .all()
    )
    return [{"button_id": r.button_id or "(sin botón)", "clicks": int(r.cnt or 0)} for r in rows]


def _devices_rows(db: Session, cid: uuid.UUID) -> list[dict[str, Any]]:
    open_rows = (
        db.query(
            models.EmailOpen.recipient_id,
            models.EmailOpen.opened_at,
            models.EmailOpen.user_agent,
            models.EmailOpen.device_category,
        )
        .filter(models.EmailOpen.campaign_id == cid)
        .order_by(
            models.EmailOpen.recipient_id,
            desc(models.EmailOpen.opened_at),
        )
        .distinct(models.EmailOpen.recipient_id)
        .all()
    )
    click_rows = (
        db.query(
            models.EmailClick.recipient_id,
            models.EmailClick.clicked_at,
            models.EmailClick.user_agent,
            models.EmailClick.device_category,
        )
        .filter(models.EmailClick.campaign_id == cid)
        .order_by(
            models.EmailClick.recipient_id,
            desc(models.EmailClick.clicked_at),
        )
        .distinct(models.EmailClick.recipient_id)
        .all()
    )

    best: dict[str, tuple[datetime, str]] = {}
    for r in open_rows:
        cat = _device_for_report_row(r.device_category, r.user_agent)
        ts = r.opened_at
        rid = r.recipient_id
        prev = best.get(rid)
        if prev is None or ts >= prev[0]:
            best[rid] = (ts, cat)

    for r in click_rows:
        cat = _device_for_report_row(r.device_category, r.user_agent)
        ts = r.clicked_at
        rid = r.recipient_id
        prev = best.get(rid)
        if prev is None or ts >= prev[0]:
            best[rid] = (ts, cat)

    counts: dict[str, int] = {}
    for _, cat in best.values():
        counts[cat] = counts.get(cat, 0) + 1
    return [{"device": k, "count": v} for k, v in sorted(counts.items(), key=lambda x: -x[1])]


def _brevo_compare_safe(db: Session, cid: uuid.UUID, campaign: models.Campaign) -> dict[str, Any] | None:
    brevo_api_key = os.getenv("BREVO_HTTP_API_KEY") or os.getenv("BREVO_API_KEY")
    if not brevo_api_key:
        return None

    rows_opens = (
        db.query(
            models.EmailOpen.recipient_id,
            func.count(models.EmailOpen.id).label("opens"),
        )
        .filter(models.EmailOpen.campaign_id == cid)
        .group_by(models.EmailOpen.recipient_id)
        .all()
    )
    internal_total_opens = sum(int(r.opens or 0) for r in rows_opens)
    internal_unique_opens = sum(1 for r in rows_opens if int(r.opens or 0) > 0)
    internal_total_clicks = (
        db.query(func.count(models.EmailClick.id))
        .filter(models.EmailClick.campaign_id == cid)
        .scalar()
        or 0
    )

    try:
        resp = httpx.get(
            "https://api.brevo.com/v3/smtp/statistics/aggregatedReport",
            headers={"accept": "application/json", "api-key": brevo_api_key},
            params={"tag": str(campaign.id), "days": 31},
            timeout=30.0,
        )
    except httpx.HTTPError:
        return None

    if resp.status_code >= 400:
        return None

    data = resp.json()
    brevo_total_opens = int(data.get("opens") or 0)
    brevo_unique_opens = int(data.get("uniqueOpens") or 0)
    brevo_total_clicks = int(data.get("clicks") or 0)
    brevo_delivered = int(data.get("delivered") or 0)
    brevo_hard_bounces = int(data.get("hardBounces") or 0)
    brevo_soft_bounces = int(data.get("softBounces") or 0)
    brevo_spam_reports = int(data.get("spamReports") or 0)

    return {
        "metric": [
            "Aperturas únicas (destinatarios)",
            "Total eventos apertura",
            "Total clics",
            "Entregados (solo Brevo)",
            "Rebotes duros (solo Brevo)",
            "Rebotes blandos (solo Brevo)",
            "Spam reportados (solo Brevo)",
        ],
        "interno": [
            internal_unique_opens,
            internal_total_opens,
            internal_total_clicks,
            "",
            "",
            "",
            "",
        ],
        "brevo": [
            brevo_unique_opens,
            brevo_total_opens,
            brevo_total_clicks,
            brevo_delivered,
            brevo_hard_bounces,
            brevo_soft_bounces,
            brevo_spam_reports,
        ],
    }


def _radar_rows(
    total_recipients: int,
    total_opens: int,
    total_clicks: int,
    unique_opens: int,
    unique_clickers: int,
    sender_name: str,
) -> list[dict[str, Any]]:
    escala = max(total_recipients, total_opens, total_clicks, 1)
    norm = lambda v: min(100.0, (float(v) / float(escala)) * 100.0)
    engagement = ((unique_opens + unique_clickers) / total_recipients * 100.0) if total_recipients else 0.0
    return [
        {"eje": "Remitente", "puntuacion_0_100": None, "valor_bruto": None, "detalle": sender_name},
        {
            "eje": "Aperturas",
            "puntuacion_0_100": round(norm(total_opens), 2),
            "valor_bruto": total_opens,
            "detalle": f"{unique_opens} destinatarios únicos",
        },
        {
            "eje": "Clics",
            "puntuacion_0_100": round(norm(total_clicks), 2),
            "valor_bruto": total_clicks,
            "detalle": f"{unique_clickers} destinatarios únicos",
        },
        {
            "eje": "Engagement",
            "puntuacion_0_100": round(min(100.0, engagement), 2),
            "valor_bruto": round(engagement, 2),
            "detalle": "únicos con apertura o clic / destinatarios",
        },
    ]


def _build_actividad_rows(
    recipients: list[dict[str, Any]], clicks_map: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in recipients:
        rid = r["recipient_id"]
        c = clicks_map.get(rid, {"click_count": 0, "last_click_at": None})
        out.append(
            {
                "recipient_id": rid,
                "email": r["email"],
                "aperturas": r["open_count"],
                "ultima_apertura": r["last_open_at"],
                "clics": c["click_count"],
                "ultimo_clic": c["last_click_at"],
            }
        )
    return out


def _sheet_resumen(wb: Workbook, campaign: models.Campaign, metrics: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Resumen"[:31]
    ws.append(["Campo", "Valor"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    rows = [
        ("campaign_id", str(campaign.id)),
        ("nombre", campaign.name or ""),
        ("asunto", campaign.subject or ""),
        ("estado", campaign.status or ""),
        ("programada_utc", campaign.scheduled_at.isoformat() if campaign.scheduled_at else ""),
        ("zona_horaria", campaign.timezone or ""),
        ("remitente", metrics.get("sender_name") or ""),
        ("total_destinatarios", metrics["total_recipients"]),
        ("destinatarios_con_apertura", metrics["unique_opens"]),
        ("total_eventos_apertura", metrics["total_opens"]),
        ("total_clics", metrics["total_clicks"]),
        ("destinatarios_con_clic", metrics["unique_clickers"]),
        ("tasa_apertura_aprox", metrics["open_rate"]),
        ("ctr_aprox", metrics["ctr"]),
        ("exportado_utc", datetime.now(timezone.utc).isoformat()),
    ]
    for k, v in rows:
        ws.append([k, v])


def _append_table_sheet(wb: Workbook, title: str, headers: list[str], data_rows: list[list[Any]]) -> None:
    name = title[:31]
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for row in data_rows:
        ws.append(row)


def _workbook_for_sections(
    campaign: models.Campaign,
    section: Section,
    metrics: dict[str, Any],
    actividad: list[dict[str, Any]],
    dispositivos: list[dict[str, Any]],
    botones: list[dict[str, Any]],
    brevo: dict[str, Any] | None,
    radar: list[dict[str, Any]],
) -> Workbook:
    wb = Workbook()
    sections = (
        ["resumen", "actividad_destinatarios", "dispositivos", "clics_boton", "brevo_interno", "remitente"]
        if section == "full"
        else [section]
    )
    first = True
    if "resumen" in sections:
        if first:
            _sheet_resumen(wb, campaign, metrics)
            first = False
        else:
            wb2 = Workbook()
            _sheet_resumen(wb2, campaign, metrics)
            ws_src = wb2.active
            new_ws = wb.create_sheet(title="Resumen"[:31])
            for row in ws_src.iter_rows(values_only=True):
                new_ws.append(list(row))
    if "actividad_destinatarios" in sections:
        if first:
            wb.remove(wb.active)
            first = False
        _append_table_sheet(
            wb,
            "Actividad_destinatarios",
            ["recipient_id", "email", "aperturas", "ultima_apertura", "clics", "ultimo_clic"],
            [
                [
                    a["recipient_id"],
                    a["email"],
                    a["aperturas"],
                    a["ultima_apertura"],
                    a["clics"],
                    a["ultimo_clic"],
                ]
                for a in actividad
            ],
        )
    if "dispositivos" in sections:
        if first:
            wb.remove(wb.active)
            first = False
        _append_table_sheet(
            wb,
            "Dispositivos",
            ["categoria_dispositivo", "destinatarios"],
            [[d["device"], d["count"]] for d in dispositivos],
        )
    if "clics_boton" in sections:
        if first:
            wb.remove(wb.active)
            first = False
        _append_table_sheet(
            wb,
            "Clics_por_boton",
            ["boton_id", "clics"],
            [[b["button_id"], b["clicks"]] for b in botones],
        )
    if "brevo_interno" in sections:
        if first:
            wb.remove(wb.active)
            first = False
        if brevo:
            _append_table_sheet(
                wb,
                "Brevo_vs_interno",
                ["metrica", "interno", "brevo"],
                [
                    [brevo["metric"][i], brevo["interno"][i], brevo["brevo"][i]]
                    for i in range(len(brevo["metric"]))
                ],
            )
        else:
            ws = wb.create_sheet(title="Brevo_vs_interno"[:31])
            ws.append(["nota"])
            ws.append([BREVO_EXPORT_NOTE])
    if "remitente" in sections:
        if first:
            wb.remove(wb.active)
            first = False
        _append_table_sheet(
            wb,
            "Remitente_radar",
            ["eje", "puntuacion_0_100", "valor_bruto", "detalle"],
            [
                [r["eje"], r["puntuacion_0_100"], r["valor_bruto"], r["detalle"]]
                for r in radar
            ],
        )
    return wb


def _csv_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\n")
    w.writerow(headers)
    for row in rows:
        w.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def build_campaign_detail_export(
    db: Session,
    campaign_id: str,
    fmt: Literal["xlsx", "csv"],
    section: Section,
) -> tuple[bytes, str, str]:
    """
    Returns (body, content_type, filename).
    For fmt=csv and section=full, body is zip bytes; filename ends with .zip.
    """
    try:
        cid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise ValueError("Campaña no encontrada") from exc

    campaign = _fetch_campaign(db, cid)
    if not campaign:
        raise ValueError("Campaña no encontrada")

    sender_display = campaign.sender.full_name if campaign.sender else ""

    total_recipients, unique_opens, total_opens, recipients = _opens_recipients(db, cid)
    clicks_map = _clicks_by_recipient_map(db, cid)
    total_clicks = sum(x["click_count"] for x in clicks_map.values())
    unique_clickers = sum(1 for x in clicks_map.values() if x["click_count"] > 0)
    open_rate = (unique_opens / total_recipients * 100.0) if total_recipients else 0.0
    ctr = (unique_clickers / total_recipients * 100.0) if total_recipients else 0.0

    metrics = {
        "sender_name": sender_display,
        "total_recipients": total_recipients,
        "unique_opens": unique_opens,
        "total_opens": total_opens,
        "total_clicks": total_clicks,
        "unique_clickers": unique_clickers,
        "open_rate": round(open_rate, 4),
        "ctr": round(ctr, 4),
    }
    actividad = _build_actividad_rows(recipients, clicks_map)
    dispositivos = _devices_rows(db, cid)
    botones = _clicks_by_button_rows(db, cid)
    brevo = _brevo_compare_safe(db, cid, campaign)
    radar = _radar_rows(
        total_recipients, total_opens, total_clicks, unique_opens, unique_clickers, sender_display or "—"
    )

    base = _sanitize_filename(campaign.name or campaign_id)

    if fmt == "xlsx":
        wb = _workbook_for_sections(campaign, section, metrics, actividad, dispositivos, botones, brevo, radar)
        bio = io.BytesIO()
        wb.save(bio)
        body = bio.getvalue()
        suffix = "completo" if section == "full" else section
        filename = f"{base}_resultados_{suffix}.xlsx"
        return body, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    if section == "full":
        zbuf = io.BytesIO()
        with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(
                "resumen.csv",
                _csv_bytes(
                    ["campo", "valor"],
                    [
                        ["campaign_id", str(campaign.id)],
                        ["nombre", campaign.name or ""],
                        ["asunto", campaign.subject or ""],
                        ["estado", campaign.status or ""],
                        ["programada_utc", campaign.scheduled_at.isoformat() if campaign.scheduled_at else ""],
                        ["zona_horaria", campaign.timezone or ""],
                        ["remitente", metrics["sender_name"]],
                        ["total_destinatarios", metrics["total_recipients"]],
                        ["destinatarios_con_apertura", metrics["unique_opens"]],
                        ["total_eventos_apertura", metrics["total_opens"]],
                        ["total_clics", metrics["total_clicks"]],
                        ["destinatarios_con_clic", metrics["unique_clickers"]],
                        ["tasa_apertura_aprox", metrics["open_rate"]],
                        ["ctr_aprox", metrics["ctr"]],
                        ["exportado_utc", datetime.now(timezone.utc).isoformat()],
                    ],
                ),
            )
            zf.writestr(
                "actividad_destinatarios.csv",
                _csv_bytes(
                    ["recipient_id", "email", "aperturas", "ultima_apertura", "clics", "ultimo_clic"],
                    [
                        [a["recipient_id"], a["email"], a["aperturas"], a["ultima_apertura"], a["clics"], a["ultimo_clic"]]
                        for a in actividad
                    ],
                ),
            )
            zf.writestr(
                "dispositivos.csv",
                _csv_bytes(
                    ["categoria_dispositivo", "destinatarios"],
                    [[d["device"], d["count"]] for d in dispositivos],
                ),
            )
            zf.writestr(
                "clics_por_boton.csv",
                _csv_bytes(
                    ["boton_id", "clics"],
                    [[b["button_id"], b["clicks"]] for b in botones],
                ),
            )
            if brevo:
                zf.writestr(
                    "brevo_vs_interno.csv",
                    _csv_bytes(
                        ["metrica", "interno", "brevo"],
                        [
                            [brevo["metric"][i], brevo["interno"][i], brevo["brevo"][i]]
                            for i in range(len(brevo["metric"]))
                        ],
                    ),
                )
            else:
                zf.writestr("brevo_vs_interno.csv", _csv_bytes(["nota"], [[BREVO_EXPORT_NOTE]]))
            zf.writestr(
                "remitente_radar.csv",
                _csv_bytes(
                    ["eje", "puntuacion_0_100", "valor_bruto", "detalle"],
                    [[r["eje"], r["puntuacion_0_100"], r["valor_bruto"], r["detalle"]] for r in radar],
                ),
            )
        filename = f"{base}_resultados_completo_csv.zip"
        return zbuf.getvalue(), "application/zip", filename

    if section == "resumen":
        body = _csv_bytes(
            ["campo", "valor"],
            [
                ["campaign_id", str(campaign.id)],
                ["nombre", campaign.name or ""],
                ["asunto", campaign.subject or ""],
                ["estado", campaign.status or ""],
                ["programada_utc", campaign.scheduled_at.isoformat() if campaign.scheduled_at else ""],
                ["zona_horaria", campaign.timezone or ""],
                ["remitente", metrics["sender_name"]],
                ["total_destinatarios", metrics["total_recipients"]],
                ["destinatarios_con_apertura", metrics["unique_opens"]],
                ["total_eventos_apertura", metrics["total_opens"]],
                ["total_clics", metrics["total_clicks"]],
                ["destinatarios_con_clic", metrics["unique_clickers"]],
                ["tasa_apertura_aprox", metrics["open_rate"]],
                ["ctr_aprox", metrics["ctr"]],
                ["exportado_utc", datetime.now(timezone.utc).isoformat()],
            ],
        )
        return body, "text/csv; charset=utf-8", f"{base}_resultados_resumen.csv"
    if section == "actividad_destinatarios":
        body = _csv_bytes(
            ["recipient_id", "email", "aperturas", "ultima_apertura", "clics", "ultimo_clic"],
            [
                [a["recipient_id"], a["email"], a["aperturas"], a["ultima_apertura"], a["clics"], a["ultimo_clic"]]
                for a in actividad
            ],
        )
        return body, "text/csv; charset=utf-8", f"{base}_resultados_actividad.csv"
    if section == "dispositivos":
        body = _csv_bytes(
            ["categoria_dispositivo", "destinatarios"],
            [[d["device"], d["count"]] for d in dispositivos],
        )
        return body, "text/csv; charset=utf-8", f"{base}_resultados_dispositivos.csv"
    if section == "clics_boton":
        body = _csv_bytes(
            ["boton_id", "clics"],
            [[b["button_id"], b["clicks"]] for b in botones],
        )
        return body, "text/csv; charset=utf-8", f"{base}_resultados_clics_boton.csv"
    if section == "brevo_interno":
        if brevo:
            body = _csv_bytes(
                ["metrica", "interno", "brevo"],
                [
                    [brevo["metric"][i], brevo["interno"][i], brevo["brevo"][i]]
                    for i in range(len(brevo["metric"]))
                ],
            )
        else:
            body = _csv_bytes(["nota"], [[BREVO_EXPORT_NOTE]])
        return body, "text/csv; charset=utf-8", f"{base}_resultados_brevo.csv"
    if section == "remitente":
        body = _csv_bytes(
            ["eje", "puntuacion_0_100", "valor_bruto", "detalle"],
            [[r["eje"], r["puntuacion_0_100"], r["valor_bruto"], r["detalle"]] for r in radar],
        )
        return body, "text/csv; charset=utf-8", f"{base}_resultados_remitente.csv"
    raise ValueError("section inválida")
