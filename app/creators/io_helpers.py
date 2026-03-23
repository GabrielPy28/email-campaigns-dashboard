"""Parseo CSV / XLSX para altas masivas de creadores."""

from __future__ import annotations

import csv
import io
from typing import Any, Iterator
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.creators.schemas import AccountProfileIn
from app.db import models


def normalize_csv_header(h: str) -> str:
    return h.strip().lower().replace(" ", "_").replace("-", "_")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _legacy_row_aliases(row: dict[str, Any]) -> dict[str, Any]:
    """Cabeceras antiguas (plantillas / campañas) -> campos actuales de `creators`."""
    r = dict(row)

    def fill(dst: str, src: str) -> None:
        if src not in r:
            return
        sv = _cell_str(r.get(src))
        if not sv:
            return
        if dst not in r or not _cell_str(r.get(dst)):
            r[dst] = r[src]

    fill("instagram_username", "handle_instagram")
    fill("tiktok_username", "handle_tiktok")
    fill("youtube_channel_url", "youtube_url")
    fill("category", "vertical")
    return r


def row_to_creator_kwargs(row: dict[str, Any]) -> dict[str, Any]:
    r = _legacy_row_aliases(row)
    out: dict[str, Any] = {}
    raw_email = r.get("email")
    email = str(raw_email).strip() if raw_email is not None else ""
    if not email:
        return out
    out["email"] = email
    for k in (
        "first_name",
        "last_name",
        "full_name",
        "picture",
        "username",
        "instagram_url",
        "tiktok_url",
        "youtube_channel_url",
        "tiktok_username",
        "instagram_username",
        "youtube_channel",
        "category",
        "facebook_page",
        "personalized_paragraph",
        "main_platform",
    ):
        if k in r and r[k] is not None and _cell_str(r[k]):
            out[k] = _cell_str(r[k])
    max_from_col: int | None = None
    if "max_followers" in r and r["max_followers"] not in (None, ""):
        try:
            max_from_col = int(float(str(r["max_followers"]).strip()))
        except ValueError:
            max_from_col = None
    if max_from_col is not None:
        out["max_followers"] = max_from_col
    else:
        follower_vals: list[int] = []
        for col in ("instagram_followers", "tiktok_followers", "youtube_subscribers"):
            if col in r and r[col] not in (None, ""):
                try:
                    follower_vals.append(int(float(str(r[col]).strip())))
                except ValueError:
                    pass
        if follower_vals:
            out["max_followers"] = max(follower_vals)
    st = _cell_str(r.get("status")).lower()
    if st in ("activo", "inactivo"):
        out["status"] = st
    return out


def platform_slug_to_id_map(db: Session) -> dict[str, UUID]:
    """Slug normalizado (como cabecera CSV) -> id de `platforms`."""
    rows = db.query(models.Platform).all()
    return {normalize_csv_header(p.nombre): p.id for p in rows}


def _parse_truthy(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = _cell_str(v).lower()
    return s in ("1", "true", "yes", "y", "x", "si", "sí", "verdadero")


def _opt_cell_str(r: dict[str, Any], key: str) -> str | None:
    if key not in r or r[key] is None:
        return None
    t = _cell_str(r[key])
    return t or None


def account_profiles_from_row(r: dict[str, Any], slug_to_id: dict[str, UUID]) -> list[AccountProfileIn]:
    """
    Columnas `{slug}_username`, `{slug}_url`, `{slug}_followers`, etc.
    slug = nombre de plataforma normalizado (p. ej. instagram, tiktok, youtube).
    """
    profiles: list[AccountProfileIn] = []
    for slug, platform_id in slug_to_id.items():
        prefix = f"{slug}_"
        username = _cell_str(r.get(prefix + "username"))
        url = _cell_str(r.get(prefix + "url"))
        if slug == "instagram":
            username = username or _cell_str(r.get("instagram_username"))
            url = url or _cell_str(r.get("instagram_url"))
        elif slug == "tiktok":
            username = username or _cell_str(r.get("tiktok_username"))
            url = url or _cell_str(r.get("tiktok_url"))
        elif slug == "youtube":
            username = username or _cell_str(r.get("youtube_channel"))
            url = url or _cell_str(r.get("youtube_channel_url"))

        if not username and not url:
            continue

        fc = 0
        follower_keys = [prefix + "followers", prefix + "followers_count"]
        if slug == "youtube":
            follower_keys.append(prefix + "subscribers")
        for fk in follower_keys:
            if fk in r and r[fk] not in (None, ""):
                try:
                    fc = int(float(str(r[fk]).strip()))
                    break
                except ValueError:
                    pass

        pc = 0
        fk_pc = prefix + "post_count"
        if fk_pc in r and r[fk_pc] not in (None, ""):
            try:
                pc = int(float(str(r[fk_pc]).strip()))
            except ValueError:
                pass

        cats: list[str] = []
        fk_cat = prefix + "category"
        if fk_cat in r and _cell_str(r.get(fk_cat)):
            parts = [p.strip() for p in _cell_str(r[fk_cat]).split(",") if p.strip()]
            cats = parts[:3]

        verified = False
        for vk in (prefix + "verified", prefix + "is_verified"):
            if vk in r and r[vk] not in (None, ""):
                verified = _parse_truthy(r[vk])
                break

        profiles.append(
            AccountProfileIn(
                platform_id=platform_id,
                username=username or None,
                url=url or None,
                picture=_opt_cell_str(r, prefix + "picture"),
                bio=_opt_cell_str(r, prefix + "bio"),
                followers_count=max(fc, 0),
                post_count=max(pc, 0),
                category=cats,
                is_verified=verified,
            )
        )
    return profiles


def row_to_creator_bulk_kwargs(row: dict[str, Any], slug_to_id: dict[str, UUID]) -> dict[str, Any]:
    """Kwargs para `CreatorCreate`: campos planos + `account_profiles` si hay cuentas por plataforma."""
    base = row_to_creator_kwargs(row)
    if not base.get("email"):
        return base
    r = _legacy_row_aliases(dict(row))
    profiles = account_profiles_from_row(r, slug_to_id)
    if profiles:
        base["account_profiles"] = profiles
    return base


def bulk_rows_carry_email(
    rows: list[tuple[int, dict[str, Any]]],
) -> tuple[list[tuple[int, str]], list[tuple[int, str, dict[str, Any]]]]:
    """
    Hereda `email` de la fila anterior cuando la celda va vacía (misma persona, varias filas).
    Devuelve (omitidas_con_motivo, filas_con_email asignado como (line_no, email_lower, norm)).
    """
    skipped: list[tuple[int, str]] = []
    out: list[tuple[int, str, dict[str, Any]]] = []
    last_email = ""
    for line_no, norm in rows:
        em = _cell_str(norm.get("email"))
        if em:
            last_email = em
        if not last_email:
            skipped.append((line_no, "email vacío y sin fila previa con email"))
            continue
        n2 = dict(norm)
        n2["email"] = last_email
        out.append((line_no, last_email.strip().lower(), n2))
    return skipped, out


def merge_norm_dicts(norms: list[dict[str, Any]]) -> dict[str, Any]:
    """Último valor no vacío gana (útil al fusionar varias filas del mismo creador)."""
    merged: dict[str, Any] = {}
    for norm in norms:
        for k, v in norm.items():
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            merged[k] = v
    return merged


def group_bulk_rows_by_email(
    carried: list[tuple[int, str, dict[str, Any]]],
) -> list[tuple[list[int], dict[str, Any]]]:
    """Agrupa por email (insensible a mayúsculas) y fusiona columnas."""
    order: list[str] = []
    buckets: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for line_no, em_lower, norm in carried:
        if em_lower not in buckets:
            order.append(em_lower)
            buckets[em_lower] = []
        buckets[em_lower].append((line_no, norm))
    groups: list[tuple[list[int], dict[str, Any]]] = []
    for em_lower in order:
        chunk = buckets[em_lower]
        line_nos = [ln for ln, _ in chunk]
        norms = [nm for _, nm in chunk]
        merged = merge_norm_dicts(norms)
        merged["email"] = chunk[0][1].get("email", "").strip() or merged.get("email", "")
        groups.append((line_nos, merged))
    return groups


def normalize_row_keys(row: dict[str, Any | None]) -> dict[str, Any]:
    norm: dict[str, Any] = {}
    for raw_key, val in row.items():
        if raw_key is None:
            continue
        nk = normalize_csv_header(str(raw_key))
        nk = nk.replace("email_address", "email")
        norm[nk] = val
    return norm


def iter_csv_dict_rows(raw: bytes) -> Iterator[tuple[int, dict[str, Any]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return
    for i, row in enumerate(reader, start=2):
        yield i, normalize_row_keys(dict(row))


def iter_xlsx_dict_rows(raw: bytes) -> Iterator[tuple[int, dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return
        headers = [
            normalize_csv_header(str(h)) if h is not None else "" for h in header_row
        ]
        if "email" not in headers and headers:
            headers = [h.replace("email_address", "email") for h in headers]
        for line_no, data_row in enumerate(rows_iter, start=2):
            row: dict[str, Any] = {}
            for idx, cell in enumerate(data_row):
                if idx < len(headers) and headers[idx]:
                    row[headers[idx]] = cell
            yield line_no, row
    finally:
        wb.close()
