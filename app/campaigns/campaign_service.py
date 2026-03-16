import io
import random
import uuid
from datetime import datetime as dt, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook

from app.db.session import get_db
from app.db import models
from app.campaigns.models import (
    CampaignCreate,
    CampaignUpdate,
    CampaignRead,
    CampaignListRow,
    RecipientRead,
)
from app.core.security import get_current_user


campaigns_router = APIRouter(prefix="/campaigns", tags=["campaigns"])


def _recipient_to_read(rec: models.CampaignRecipient) -> RecipientRead:
    return RecipientRead(
        id=rec.recipient_id,
        email=rec.email,
        nombre=rec.nombre,
        username=rec.username,
        extra=rec.extra_data or {},
    )


def _campaign_to_read(c: models.Campaign) -> CampaignRead:
    sender_ids = [str(s.id) for s in c.senders] if c.senders else [str(c.sender_id)]
    recipients = [_recipient_to_read(r) for r in c.recipients] if c.recipients else []
    return CampaignRead(
        id=str(c.id),
        name=c.name,
        subject=c.subject,
        preheader=c.preheader,
        template_id=str(c.template_id),
        sender_id=str(c.sender_id),
        sender_name=c.sender.full_name,
        sender_ids=sender_ids,
        scheduled_at=c.scheduled_at,
        timezone=c.timezone,
        wait_min_seconds=c.wait_min_seconds,
        wait_max_seconds=c.wait_max_seconds,
        status=c.status,
        created_by=c.created_by,
        created_at=c.created_at,
        recipients=recipients,
    )


@campaigns_router.post("/", response_model=CampaignRead, status_code=status.HTTP_201_CREATED)
def create_campaign(
    payload: CampaignCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    if payload.wait_max_seconds < payload.wait_min_seconds:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="wait_max_seconds must be >= wait_min_seconds",
        )

    template_id = payload.template_id
    template = db.query(models.Template).filter(models.Template.id == template_id).first()
    if not template:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Plantilla no encontrada. Registre la plantilla en POST /templates/.",
        )

    sender_uuids = list(payload.sender_ids)
    senders_in_db = db.query(models.Sender).filter(models.Sender.id.in_(sender_uuids)).all()
    if len(senders_in_db) != len(sender_uuids):
        found = {s.id for s in senders_in_db}
        missing = [str(s) for s in sender_uuids if s not in found]
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Remitente(s) no encontrado(s): {', '.join(missing)}. Registre en POST /senders/.",
        )

    primary_sender_id = sender_uuids[0]
    created_by = current_user.get("name") or current_user.get("email") or ""

    db_campaign = models.Campaign(
        name=payload.name,
        subject=payload.subject,
        preheader=payload.preheader,
        template_id=template_id,
        sender_id=primary_sender_id,
        scheduled_at=payload.scheduled_at,
        timezone=payload.timezone,
        wait_min_seconds=payload.wait_min_seconds,
        wait_max_seconds=payload.wait_max_seconds,
        created_by=created_by,
    )
    db.add(db_campaign)
    db.flush()

    for sid in sender_uuids:
        db.execute(models.campaign_senders.insert().values(campaign_id=db_campaign.id, sender_id=sid))

    for r in payload.recipients:
        required, extra = r.to_required_and_extra()
        db.add(models.CampaignRecipient(
            campaign_id=db_campaign.id,
            recipient_id=str(required["id"]),
            email=required["email"],
            nombre=required["nombre"],
            username=required["username"],
            extra_data=extra,
        ))

    db.commit()
    db.refresh(db_campaign)
    return _campaign_to_read(db_campaign)


@campaigns_router.put("/{campaign_id}", response_model=CampaignRead)
def update_campaign(
    campaign_id: str,
    payload: CampaignUpdate,
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    try:
        cid = uuid.UUID(campaign_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")

    campaign = db.query(models.Campaign).filter(models.Campaign.id == cid).first()
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")

    # No modificamos destinatarios ni métricas; solo metadatos básicos.
    if payload.name is not None:
        campaign.name = payload.name
    if payload.subject is not None:
        campaign.subject = payload.subject
    if payload.preheader is not None:
        campaign.preheader = payload.preheader
    if payload.template_id is not None:
        campaign.template_id = payload.template_id
    if payload.scheduled_at is not None:
        campaign.scheduled_at = payload.scheduled_at
    if payload.timezone is not None:
        campaign.timezone = payload.timezone
    if payload.wait_min_seconds is not None:
        campaign.wait_min_seconds = payload.wait_min_seconds
    if payload.wait_max_seconds is not None:
        campaign.wait_max_seconds = payload.wait_max_seconds

    db.commit()
    db.refresh(campaign)
    return _campaign_to_read(campaign)


@campaigns_router.post("/{campaign_id}/prepare-send", response_model=CampaignRead)
def prepare_campaign_send(
    campaign_id: str,
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    """
    Paso de pre-envío:
    - Asigna sender_id a cada destinatario (distribución equitativa sobre sender_ids).
    - Calcula scheduled_send_time por destinatario usando wait_min_seconds/wait_max_seconds.
    - Marca status='pending' en campaign_recipients.
    """
    try:
        cid = uuid.UUID(campaign_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")

    campaign = (
        db.query(models.Campaign)
        .options(
            joinedload(models.Campaign.senders),
            joinedload(models.Campaign.recipients),
        )
        .filter(models.Campaign.id == cid)
        .first()
    )
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")

    if not campaign.recipients:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La campaña no tiene destinatarios registrados.",
        )

    sender_ids = [s.id for s in campaign.senders] if campaign.senders else [campaign.sender_id]
    if not sender_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La campaña no tiene remitentes asociados.",
        )

    # Distribución equitativa de senders sobre recipients
    recipients_sorted = sorted(campaign.recipients, key=lambda r: (r.recipient_id, r.email))
    for idx, rec in enumerate(recipients_sorted):
        rec.sender_id = sender_ids[idx % len(sender_ids)]
        rec.status = "pending"

    # Programación de tiempos de envío por destinatario
    current_time = campaign.scheduled_at
    if current_time is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La campaña no tiene fecha/hora de envío (scheduled_at).",
        )

    for idx, rec in enumerate(recipients_sorted):
        if idx == 0:
            rec.scheduled_send_time = current_time
        else:
            delta_seconds = random.randint(
                campaign.wait_min_seconds,
                campaign.wait_max_seconds,
            )
            current_time = current_time + timedelta(seconds=delta_seconds)
            rec.scheduled_send_time = current_time

    db.commit()
    db.refresh(campaign)
    return _campaign_to_read(campaign)


def _campaign_list_row(
    c: models.Campaign,
    num_recipients: int,
    sent_at: Optional[dt],
    total_opens: int,
    total_clicks: int,
) -> CampaignListRow:
    open_rate = (total_opens / num_recipients * 100) if num_recipients else 0.0
    click_rate = (total_clicks / num_recipients * 100) if num_recipients else 0.0
    return CampaignListRow(
        id=str(c.id),
        name=c.name,
        sender_name=c.sender.full_name if c.sender else "",
        subject=c.subject,
        preheader=c.preheader,
        template_name=c.template.name if c.template else None,
        num_recipients=num_recipients,
        sent_at=sent_at,
        total_opens=total_opens,
        total_clicks=total_clicks,
        open_rate_pct=round(open_rate, 1),
        click_rate_pct=round(click_rate, 1),
        status=c.status,
        created_by=c.created_by,
        created_at=c.created_at,
        scheduled_at=c.scheduled_at,
    )


@campaigns_router.get("/", response_model=List[CampaignListRow])
def list_campaigns(
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
    scheduled_at_from: Optional[str] = Query(None, description="Fecha envío desde (YYYY-MM-DD)"),
    scheduled_at_to: Optional[str] = Query(None, description="Fecha envío hasta (YYYY-MM-DD)"),
    sender_id: Optional[str] = Query(None, description="UUID del sender"),
    sender_name: Optional[str] = Query(None, description="Nombre del sender (contiene)"),
    subject: Optional[str] = Query(None, description="Asunto (contiene)"),
    preheader: Optional[str] = Query(None, description="Preheader (contiene)"),
    created_by: Optional[str] = Query(None, description="Creado por (contiene)"),
    created_at_from: Optional[str] = Query(None, description="Creado desde (YYYY-MM-DD)"),
    created_at_to: Optional[str] = Query(None, description="Creado hasta (YYYY-MM-DD)"),
    template_id: Optional[str] = Query(None, description="UUID de la plantilla"),
    template_name: Optional[str] = Query(None, description="Nombre plantilla (contiene)"),
    campaign_id: Optional[str] = Query(None, description="UUID de la campaña"),
    campaign_name: Optional[str] = Query(None, description="Nombre campaña (contiene)"),
):
    q = (
        db.query(models.Campaign)
        .options(
            joinedload(models.Campaign.sender),
            joinedload(models.Campaign.template),
        )
        .order_by(models.Campaign.created_at.desc())
    )
    if scheduled_at_from:
        try:
            q = q.filter(models.Campaign.scheduled_at >= dt.fromisoformat(scheduled_at_from.replace("Z", "+00:00")))
        except ValueError:
            pass
    if scheduled_at_to:
        try:
            q = q.filter(models.Campaign.scheduled_at <= dt.fromisoformat(scheduled_at_to.replace("Z", "+00:00")))
        except ValueError:
            pass
    if sender_id:
        try:
            sid = uuid.UUID(sender_id)
            q = q.filter(models.Campaign.sender_id == sid)
        except ValueError:
            pass
    if sender_name:
        q = q.join(models.Sender, models.Campaign.sender_id == models.Sender.id).filter(
            models.Sender.full_name.ilike(f"%{sender_name}%")
        )
    if subject:
        q = q.filter(or_(models.Campaign.subject.is_(None), models.Campaign.subject.ilike(f"%{subject}%")))
    if preheader:
        q = q.filter(or_(models.Campaign.preheader.is_(None), models.Campaign.preheader.ilike(f"%{preheader}%")))
    if created_by:
        q = q.filter(models.Campaign.created_by.ilike(f"%{created_by}%"))
    if created_at_from:
        try:
            q = q.filter(models.Campaign.created_at >= dt.fromisoformat(created_at_from.replace("Z", "+00:00")))
        except ValueError:
            pass
    if created_at_to:
        try:
            q = q.filter(models.Campaign.created_at <= dt.fromisoformat(created_at_to.replace("Z", "+00:00")))
        except ValueError:
            pass
    if template_id:
        try:
            tid = uuid.UUID(template_id)
            q = q.filter(models.Campaign.template_id == tid)
        except ValueError:
            pass
    if template_name:
        q = q.join(models.Template, models.Campaign.template_id == models.Template.id).filter(
            or_(models.Template.name.is_(None), models.Template.name.ilike(f"%{template_name}%"))
        )
    if campaign_id:
        try:
            cid = uuid.UUID(campaign_id)
            q = q.filter(models.Campaign.id == cid)
        except ValueError:
            pass
    if campaign_name:
        q = q.filter(models.Campaign.name.ilike(f"%{campaign_name}%"))

    campaigns = q.distinct().all()
    if not campaigns:
        return []

    cids = [c.id for c in campaigns]
    # num_recipients por campaña
    rec_counts = (
        db.query(models.CampaignRecipient.campaign_id, func.count(models.CampaignRecipient.id).label("n"))
        .filter(models.CampaignRecipient.campaign_id.in_(cids))
        .group_by(models.CampaignRecipient.campaign_id)
        .all()
    )
    rec_map = {r[0]: int(r[1]) for r in rec_counts}
    # sent_at (max scheduled_send_time donde status=sent)
    sent_times = (
        db.query(models.CampaignRecipient.campaign_id, func.max(models.CampaignRecipient.scheduled_send_time).label("t"))
        .filter(models.CampaignRecipient.campaign_id.in_(cids), models.CampaignRecipient.status == "sent")
        .group_by(models.CampaignRecipient.campaign_id)
        .all()
    )
    sent_map = {s[0]: s[1] for s in sent_times}
    # total_opens
    open_counts = (
        db.query(models.EmailOpen.campaign_id, func.count(models.EmailOpen.id).label("n"))
        .filter(models.EmailOpen.campaign_id.in_(cids))
        .group_by(models.EmailOpen.campaign_id)
        .all()
    )
    open_map = {o[0]: int(o[1]) for o in open_counts}
    # total_clicks
    click_counts = (
        db.query(models.EmailClick.campaign_id, func.count(models.EmailClick.id).label("n"))
        .filter(models.EmailClick.campaign_id.in_(cids))
        .group_by(models.EmailClick.campaign_id)
        .all()
    )
    click_map = {cl[0]: int(cl[1]) for cl in click_counts}

    return [
        _campaign_list_row(
            c,
            num_recipients=rec_map.get(c.id, 0),
            sent_at=sent_map.get(c.id),
            total_opens=open_map.get(c.id, 0),
            total_clicks=click_map.get(c.id, 0),
        )
        for c in campaigns
    ]


@campaigns_router.delete("/{campaign_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_campaign(
    campaign_id: str,
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    try:
        cid = uuid.UUID(campaign_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")

    campaign = db.query(models.Campaign).filter(models.Campaign.id == cid).first()
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")

    # Eliminar asociaciones en campaign_senders explícitamente y confiar en cascade para recipients.
    db.execute(models.campaign_senders.delete().where(models.campaign_senders.c.campaign_id == cid))
    db.delete(campaign)
    db.commit()
    return None


@campaigns_router.get("/export/excel")
def export_campaigns_excel(
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
    scheduled_at_from: Optional[str] = Query(None),
    scheduled_at_to: Optional[str] = Query(None),
    sender_id: Optional[str] = Query(None),
    sender_name: Optional[str] = Query(None),
    subject: Optional[str] = Query(None),
    preheader: Optional[str] = Query(None),
    created_by: Optional[str] = Query(None),
    created_at_from: Optional[str] = Query(None),
    created_at_to: Optional[str] = Query(None),
    template_id: Optional[str] = Query(None),
    template_name: Optional[str] = Query(None),
    campaign_id: Optional[str] = Query(None),
    campaign_name: Optional[str] = Query(None),
):
    """Descarga Excel con el listado de campañas (mismos filtros que GET /campaigns/)."""
    rows = list_campaigns(
        db=db,
        _=_,
        scheduled_at_from=scheduled_at_from,
        scheduled_at_to=scheduled_at_to,
        sender_id=sender_id,
        sender_name=sender_name,
        subject=subject,
        preheader=preheader,
        created_by=created_by,
        created_at_from=created_at_from,
        created_at_to=created_at_to,
        template_id=template_id,
        template_name=template_name,
        campaign_id=campaign_id,
        campaign_name=campaign_name,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Campañas"
    headers = [
        "ID", "Nombre", "Sender", "Asunto", "Preheader", "Plantilla",
        "Num. Recipients", "Sent at", "Aperturas", "Clics", "% Aperturas", "% Clics",
        "Estado", "Creado por", "Creado at", "Scheduled at",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.name)
        ws.cell(row=row_idx, column=3, value=r.sender_name)
        ws.cell(row=row_idx, column=4, value=r.subject or "")
        ws.cell(row=row_idx, column=5, value=r.preheader or "")
        ws.cell(row=row_idx, column=6, value=r.template_name or "")
        ws.cell(row=row_idx, column=7, value=r.num_recipients)
        ws.cell(row=row_idx, column=8, value=r.sent_at.strftime("%d-%m-%Y %H:%M") if r.sent_at else "")
        ws.cell(row=row_idx, column=9, value=r.total_opens)
        ws.cell(row=row_idx, column=10, value=r.total_clicks)
        ws.cell(row=row_idx, column=11, value=r.open_rate_pct)
        ws.cell(row=row_idx, column=12, value=r.click_rate_pct)
        ws.cell(row=row_idx, column=13, value=r.status)
        ws.cell(row=row_idx, column=14, value=r.created_by)
        ws.cell(row=row_idx, column=15, value=r.created_at.strftime("%d-%m-%Y %H:%M") if r.created_at else "")
        ws.cell(row=row_idx, column=16, value=r.scheduled_at.strftime("%d-%m-%Y %H:%M") if r.scheduled_at else "")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=campaigns.xlsx"},
    )


@campaigns_router.get("/{campaign_id}", response_model=CampaignRead)
def get_campaign(
    campaign_id: str,
    db: Session = Depends(get_db),
    _: dict = Depends(get_current_user),
):
    try:
        cid = uuid.UUID(campaign_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    campaign = (
        db.query(models.Campaign)
        .options(
            joinedload(models.Campaign.sender),
            joinedload(models.Campaign.senders),
            joinedload(models.Campaign.recipients),
        )
        .filter(models.Campaign.id == cid)
        .first()
    )
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    return _campaign_to_read(campaign)
